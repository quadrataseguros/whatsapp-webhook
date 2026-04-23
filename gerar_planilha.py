#!/usr/bin/env python3
"""
Planilha de Comissões Wilson - Quadrata Seguros
Vigência: 23/04/2025 a 23/04/2026
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------------------------
# DADOS  (mês, cliente, seguradora, item, premio, coms%, com, 26%, 2%da,
#          recebido, vezes, receber, parc_num, parc_total)
# ---------------------------------------------------------------------------
MESES = {
    "abr/25": [
        ("LEIDE DAYANE DE SOUZA","SUHAI","TORO",2985.74,25,746.44,194.07,3.88,190.19,12,15.85,12,12),
        ("VALDISLEI XAVIER DO NASCIMENTO","BRASDESCO","COROLLA",4099.42,10,409.94,106.58,2.13,104.45,10,10.45,10,10),
        ("CRISTIANE DOS SANTOS COSTA SILVA","ALLIANZ","ONIX",2023.63,20,404.73,105.23,2.10,103.12,10,10.31,8,10),
        ("LEIDE DAYANE DE SOUZA","ALIRO","ECOSPORT",1786.77,20,357.35,92.91,1.86,91.05,10,9.11,8,10),
        ("LUIZ ANTONIO DE CARVALHO","AZUL","SANDERO",1669.87,20,333.97,86.83,1.74,85.10,10,8.51,6,10),
        ("JULIO CESAR ALMEIDA DE OLIVEIRA","TOKIO MARINE","I30",2089.08,20,417.82,108.63,2.17,106.46,12,8.87,6,12),
        ("SOLENIR GOMES DA SILVA MANCINI","AZUL","JETTA",4210.75,15,631.61,164.22,3.28,160.93,10,16.09,6,10),
        ("CHRISTIAN DO NASCIMENTO","TOKIO MARINE","TIGGO",2320.42,20,464.08,120.66,2.41,118.25,5,23.65,5,5),
        ("CLAUDEMAR RAIMUNDO DO NASCIMENTO","PORTO","ONIX",1839.91,20,367.98,95.68,1.91,93.76,10,9.38,5,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","CRETA",1752.74,20,350.55,91.14,1.82,89.32,12,7.44,5,12),
        ("ROBSON NASCIMENTO SANTOS","SUHAI","CITY",3329.52,15,499.43,129.85,2.60,127.25,12,10.60,5,12),
        ("JONATAS DE BRITO LUCAS","TOKIO MARINE","VERSA",2121.51,20,424.30,110.32,2.21,108.11,7,15.44,4,7),
        ("PAULO SERGIO LADUANO","TOKIO MARINE","COROLLA",1790.54,20,358.11,93.11,1.86,91.25,5,18.25,4,5),
        ("LUIZ ANTONIO DE CARVALHO","SUHAI","SPORTAGE",1640.47,20,328.09,85.30,1.71,83.60,8,10.45,4,8),
        ("RAFAEL SANTOS NASCIMENTO","AZUL","CIVIC",3336.34,20,667.27,173.49,3.47,170.02,10,17.00,3,10),
        ("ELIZABETH IZAGUIRRE GOMES","HDI","PALIO",1029.05,20,205.81,53.51,1.07,52.44,10,5.24,3,10),
        ("FRANCISCO FERNANDES DE LIMA","TOKIO MARINE","CITY",1613.57,20,322.71,83.91,1.68,82.23,12,6.85,3,12),
        ("SOLANGE SILVA DE JESUS","TOKIO MARINE","I30",1602.96,20,320.59,83.35,1.67,81.69,12,6.81,3,12),
        ("ROSANGELA DE JESUS","SUHAI","IDEA",1454.42,25,363.61,94.54,1.89,92.65,1,92.65,1,1),
    ],
    "mai/25": [],  # dados parciais – preencher manualmente
    "jun/25": [
        ("LEIDE DAYANE DE SOUZA","SUHAI","TORO",2985.74,25,746.44,194.07,3.88,190.19,12,15.85,12,12),
        ("VALDISLEI XAVIER DO NASCIMENTO","BRASDESCO","COROLLA",4099.42,10,409.94,106.58,2.13,104.45,10,10.45,10,10),
        ("CRISTIANE DOS SANTOS COSTA SILVA","ALLIANZ","ONIX",2023.63,20,404.73,105.23,2.10,103.12,10,10.31,8,10),
        ("LEIDE DAYANE DE SOUZA","ALIRO","ECOSPORT",1786.77,20,357.35,92.91,1.86,91.05,10,9.11,8,10),
        ("LUIZ ANTONIO DE CARVALHO","AZUL","SANDERO",1669.87,20,333.97,86.83,1.74,85.10,10,8.51,6,10),
        ("JULIO CESAR ALMEIDA DE OLIVEIRA","TOKIO MARINE","I30",2089.08,20,417.82,108.63,2.17,106.46,12,8.87,6,12),
        ("SOLENIR GOMES DA SILVA MANCINI","AZUL","JETTA",4210.75,15,631.61,164.22,3.28,160.93,10,16.09,6,10),
        ("CHRISTIAN DO NASCIMENTO","TOKIO MARINE","TIGGO",2320.42,20,464.08,120.66,2.41,118.25,5,23.65,5,5),
        ("CLAUDEMAR RAIMUNDO DO NASCIMENTO","PORTO","ONIX",1839.91,20,367.98,95.68,1.91,93.76,10,9.38,5,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","CRETA",1752.74,20,350.55,91.14,1.82,89.32,12,7.44,5,12),
        ("ROBSON NASCIMENTO SANTOS","SUHAI","CITY",3329.52,15,499.43,129.85,2.60,127.25,12,10.60,5,12),
        ("JONATAS DE BRITO LUCAS","TOKIO MARINE","VERSA",2121.51,20,424.30,110.32,2.21,108.11,7,15.44,4,7),
        ("LUIZ ANTONIO DE CARVALHO","SUHAI","SPORTAGE",1640.47,20,328.09,85.30,1.71,83.60,8,10.45,4,8),
        ("RAFAEL SANTOS NASCIMENTO","AZUL","CIVIC",3336.34,20,667.27,173.49,3.47,170.02,10,17.00,3,10),
        ("ELIZABETH IZAGUIRRE GOMES","HDI","PALIO",1029.05,20,205.81,53.51,1.07,52.44,10,5.24,3,10),
        ("FRANCISCO FERNANDES DE LIMA","TOKIO MARINE","CITY",1613.57,20,322.71,83.91,1.68,82.23,12,6.85,3,12),
        ("SOLANGE SILVA DE JESUS","TOKIO MARINE","I30",1602.96,20,320.59,83.35,1.67,81.69,12,6.81,3,12),
        ("ANTONIO CARLOS DOS SANTOS","PORTO","TUCSON",2201.31,18,396.24,103.02,2.06,100.96,10,10.10,1,10),
        ("JOSE NILTON DOS SANTOS","PORTO","HR-V",2256.38,20,451.28,117.33,2.35,114.99,4,28.75,1,4),
        ("OTAVIANO DE JESUS","PORTO","CORSA",1854.88,18,333.88,86.81,1.74,85.07,10,8.51,1,10),
        ("LEONARDO VAZ DE FREITAS","SUHAI","TUCSON",1689.84,25,422.46,109.84,2.20,107.64,12,8.97,1,12),
        ("JURANDIR FERREIRA DOS SANTOS","TOKIO MARINE","GOL",144.04,20,28.81,7.49,1.00,6.49,1,6.49,1,1),
    ],
    "jul/25": [
        ("VALDISLEI XAVIER DO NASCIMENTO","BRASDESCO","COROLLA",4099.42,10,409.94,106.58,2.13,104.45,10,10.45,10,10),
        ("LEIDE DAYANE DE SOUZA","ALIRO","ECOSPORT",1786.77,20,357.35,92.91,1.86,91.05,10,9.11,8,10),
        ("LUIZ ANTONIO DE CARVALHO","AZUL","SANDERO",1669.87,20,333.97,86.83,1.74,85.10,10,8.51,9,10),
        ("JULIO CESAR ALMEIDA DE OLIVEIRA","TOKIO MARINE","I30",2089.08,20,417.82,108.63,2.17,106.46,12,8.87,6,12),
        ("SOLENIR GOMES DA SILVA MANCINI","AZUL","JETTA",4210.75,15,631.61,164.22,3.28,160.93,10,16.09,6,10),
        ("CHRISTIAN DO NASCIMENTO","TOKIO MARINE","TIGGO",2320.42,20,464.08,120.66,2.41,118.25,5,23.65,5,5),
        ("CLAUDEMAR RAIMUNDO DO NASCIMENTO","PORTO","ONIX",1839.91,20,367.98,95.68,1.91,93.76,10,9.38,5,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","CRETA",1752.74,20,350.55,91.14,1.82,89.32,12,7.44,5,12),
        ("ROBSON NASCIMENTO SANTOS","SUHAI","CITY",3329.52,15,499.43,129.85,2.60,127.25,12,10.60,5,12),
        ("JONATAS DE BRITO LUCAS","TOKIO MARINE","VERSA",2121.51,20,424.30,110.32,2.21,108.11,7,15.44,7,7),
        ("LUIZ ANTONIO DE CARVALHO","SUHAI","SPORTAGE",1640.47,20,328.09,85.30,1.71,83.60,8,10.45,4,8),
        ("RAFAEL SANTOS NASCIMENTO","AZUL","CIVIC",3336.34,20,667.27,173.49,3.47,170.02,10,17.00,3,10),
        ("ELIZABETH IZAGUIRRE GOMES","HDI","PALIO",1029.05,20,205.81,53.51,1.07,52.44,10,5.24,3,10),
        ("FRANCISCO FERNANDES DE LIMA","TOKIO MARINE","CITY",1613.57,20,322.71,83.91,1.68,82.23,12,6.85,3,12),
        ("SOLANGE SILVA DE JESUS","TOKIO MARINE","I30",1602.96,20,320.59,83.35,1.67,81.69,12,6.81,3,12),
        ("ROSANGELA DE JESUS MELO","SUHAI","IDEA",1454.42,25,363.61,94.54,1.89,92.65,1,92.65,2,1),
        ("ANTONIO CARLOS DOS SANTOS","PORTO","TUCSON",2201.31,18,396.24,103.02,2.06,100.96,10,10.10,2,10),
        ("JOSE NILTON DOS SANTOS","PORTO","HR-V",2256.38,20,451.28,117.33,2.35,114.99,4,28.75,2,4),
        ("OTAVIANO DE JESUS","PORTO","CORSA",1854.88,18,333.88,86.81,1.74,85.07,10,8.51,2,10),
        ("LEONARDO VAZ DE FREITAS","SUHAI","TUCSON",1689.84,25,422.46,109.84,2.20,107.64,12,8.97,2,12),
        ("VALDISLEI XAVIER DO NASCIMENTO","BRADESCO","COROLLA",4204.31,20,840.86,218.62,4.37,214.25,10,21.43,1,10),
        ("DERANILDA PEREIRA COELHO","SUHAI","ZARIFA",1553.27,20,310.65,80.77,1.62,79.15,5,15.83,1,5),
        ("IVO REIS DO NASCIMENTO","AZUL","CITY SEDAN",1199.51,20,239.90,62.37,1.25,61.13,1,61.13,1,1),
        ("ANDREA NASCIMENTO SANTANA MARTIENA","TOKIO MARINE","CRETA",1986.95,20,397.39,103.32,2.07,101.25,12,8.44,1,12),
    ],
    "ago/25": [
        ("LUIZ ANTONIO DE CARVALHO","AZUL","SANDERO",1669.87,20,333.97,86.83,1.74,85.10,10,8.51,10,10),
        ("JULIO CESAR ALMEIDA DE OLIVEIRA","TOKIO MARINE","I30",2089.08,20,417.82,108.63,2.17,106.46,12,8.87,8,12),
        ("SOLENIR GOMES DA SILVA MANCINI","AZUL","JETTA",4210.75,15,631.61,164.22,3.28,160.93,10,16.09,10,10),
        ("ROBSON NASCIMENTO SANTOS","SUHAI","CITY",3329.52,15,499.43,129.85,2.60,127.25,12,10.60,8,12),
        ("CLAUDEMAR RAIMUNDO DO NASCIMENTO","PORTO","ONIX",1839.91,20,367.98,95.68,1.91,93.76,10,9.38,7,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","CRETA",1752.74,20,350.55,91.14,1.82,89.32,12,7.44,7,12),
        ("LUIZ ANTONIO DE CARVALHO","SUHAI","SPORTAGE",1640.47,20,328.09,85.30,1.71,83.60,8,10.45,8,8),
        ("RAFAEL SANTOS NASCIMENTO","AZUL","CIVIC",3336.34,20,667.27,173.49,3.47,170.02,10,17.00,5,10),
        ("ELIZABETH IZAGUIRRE GOMES","HDI","PALIO",1029.05,20,205.81,53.51,1.07,52.44,10,5.24,5,10),
        ("FRANCISCO FERNANDES DE LIMA","TOKIO MARINE","CITY",1613.57,20,322.71,83.91,1.68,82.23,12,6.85,5,12),
        ("SOLANGE SILVA DE JESUS","TOKIO MARINE","I30",1602.96,20,320.59,83.35,1.67,81.69,12,6.81,5,12),
        ("ROSANGELA DE JESUS MELO","SUHAI","IDEA",1454.42,25,363.61,94.54,1.89,92.65,1,92.65,3,1),
        ("ANTONIO CARLOS DOS SANTOS","PORTO","TUCSON",2201.31,18,396.24,103.02,2.06,100.96,10,10.10,3,10),
        ("JOSE NILTON DOS SANTOS","PORTO","HR-V",2256.38,20,451.28,117.33,2.35,114.99,4,28.75,3,4),
        ("OTAVIANO DE JESUS","PORTO","CORSA",1854.88,18,333.88,86.81,1.74,85.07,10,8.51,3,10),
        ("LEONARDO VAZ DE FREITAS","SUHAI","TUCSON",1689.84,25,422.46,109.84,2.20,107.64,12,8.97,3,12),
        ("VALDISLEI XAVIER DO NASCIMENTO","BRADESCO","COROLLA",4204.31,20,840.86,218.62,4.37,214.25,10,21.43,2,10),
        ("DERANILDA PEREIRA COELHO","SUHAI","ZARIFA",1553.27,20,310.65,80.77,1.62,79.15,5,15.83,2,5),
        ("IVO REIS DO NASCIMENTO","AZUL","CITY SEDAN",1199.51,20,239.90,62.37,1.25,61.13,1,61.13,2,1),
        ("ANDREA NASCIMENTO SANTANA MARTIENA","TOKIO MARINE","CRETA",1986.95,20,397.39,103.32,2.07,101.25,12,8.44,2,12),
        ("CAIO CESAR DOS SANTOS","SUHAI","SIENA",3310.55,20,662.11,172.15,3.44,168.71,12,14.06,1,12),
        ("IVO REIS DO NASCIMENTO","MITSUI","FOX",2141.74,20,428.35,111.37,2.23,109.14,10,10.91,1,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","TRACKER",326.96,20,65.39,17.00,0.34,16.66,3,5.55,1,3),
    ],
    "set/25": [
        ("JULIO CESAR ALMEIDA DE OLIVEIRA","TOKIO MARINE","I30",2089.08,20,417.82,108.63,2.17,106.46,12,8.87,9,12),
        ("ROBSON NASCIMENTO SANTOS","SUHAI","CITY",3329.52,15,499.43,129.85,2.60,127.25,12,10.60,10,12),
        ("CLAUDEMAR RAIMUNDO DO NASCIMENTO","PORTO","ONIX",1839.91,20,367.98,95.68,1.91,93.76,10,9.38,10,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","CRETA",1752.74,20,350.55,91.14,1.82,89.32,12,7.44,8,12),
        ("RAFAEL SANTOS NASCIMENTO","AZUL","CIVIC",3336.34,20,667.27,173.49,3.47,170.02,10,17.00,8,10),
        ("ELIZABETH IZAGUIRRE GOMES","HDI","PALIO",1029.05,20,205.81,53.51,1.07,52.44,10,5.24,8,10),
        ("FRANCISCO FERNANDES DE LIMA","TOKIO MARINE","CITY",1613.57,20,322.71,83.91,1.68,82.23,12,6.85,10,12),
        ("SOLANGE SILVA DE JESUS","TOKIO MARINE","I30",1602.96,20,320.59,83.35,1.67,81.69,12,6.81,6,12),
        ("ROSANGELA DE JESUS MELO","SUHAI","IDEA",1454.42,25,363.61,94.54,1.89,92.65,1,92.65,4,1),
        ("ANTONIO CARLOS DOS SANTOS","PORTO","TUCSON",2201.31,18,396.24,103.02,2.06,100.96,10,10.10,4,10),
        ("JOSE NILTON DOS SANTOS","PORTO","HR-V",2256.38,20,451.28,117.33,2.35,114.99,4,28.75,4,4),
        ("OTAVIANO DE JESUS","PORTO","CORSA",1854.88,18,333.88,86.81,1.74,85.07,10,8.51,4,10),
        ("LEONARDO VAZ DE FREITAS","SUHAI","TUCSON",1689.84,25,422.46,109.84,2.20,107.64,12,8.97,4,12),
        ("VALDISLEI XAVIER DO NASCIMENTO","BRADESCO","COROLLA",4204.31,20,840.86,218.62,4.37,214.25,10,21.43,3,10),
        ("DERANILDA PEREIRA COELHO","SUHAI","ZARIFA",1553.27,20,310.65,80.77,1.62,79.15,5,15.83,3,5),
        ("IVO REIS DO NASCIMENTO","AZUL","CITY SEDAN",1199.51,20,239.90,62.37,1.25,61.13,1,61.13,3,1),
        ("ANDREA NASCIMENTO SANTANA MARTIENA","TOKIO MARINE","CRETA",1986.95,20,397.39,103.32,2.07,101.25,12,8.44,3,12),
        ("CAIO CESAR DOS SANTOS","SUHAI","SIENA",3310.55,20,662.11,172.15,3.44,168.71,12,14.06,2,12),
        ("IVO REIS DO NASCIMENTO","MITSUI","FOX",2141.74,20,428.35,111.37,2.23,109.14,10,10.91,2,10),
        ("FRANCISCO FERNANDES DE LIMA","ALIRO","TRACKER",326.96,20,65.39,17.00,0.34,16.66,3,5.55,2,3),
        ("LEIDE DAYANE DE SOUZA","ALIRO","ECOSPORT",1642.98,20,328.60,85.43,1.71,83.73,10,8.37,2,10),
    ],
    "out/25": [],
    "nov/25": [],
    "dez/25": [],
    "jan/26": [],
    "fev/26": [],
    "mar/26": [],
    "abr/26": [],
}

# ---------------------------------------------------------------------------
# ESTILOS
# ---------------------------------------------------------------------------
COR_TITULO      = "1F3864"
COR_HEADER      = "1F3864"
COR_MES         = "2E75B6"
COR_BRANCO      = "FFFFFF"
COR_CINZA_CLR   = "F2F2F2"
COR_VERDE_CLR   = "E2EFDA"
COR_VERM_CLR    = "FFE0E0"
COR_AMAR_CLR    = "FFF2CC"

BORDA_FINA = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
BORDA_MES = Border(
    left=Side(style="medium", color=COR_MES),
    right=Side(style="medium", color=COR_MES),
    top=Side(style="medium", color=COR_MES),
    bottom=Side(style="medium", color=COR_MES),
)

def fill(cor): return PatternFill("solid", fgColor=cor)
def font(bold=False, size=10, cor=None, italic=False):
    return Font(name="Calibri", bold=bold, size=size,
                color=cor or "000000", italic=italic)

# Larguras das colunas (A..R)
COL_WIDTHS = [32,14,13,11,7,12,10,8,12,7,11,5,4,5,9,14,16,10]
# Cabeçalhos de dados (colunas B em diante)
DATA_HEADERS = [
    "PRÊMIO LÍQ","SEGURADORA","ITEM","COM%","COMISSÃO",
    "26%","2% DA","RECEBIDO","VEZES","RECEBER",
    "","DE","","AJUSTE","STATUS","ADIMPLÊNCIA","OBS"
]

# ---------------------------------------------------------------------------
def build_sheet(wb):
    ws = wb.active
    ws.title = "COMISSÕES WILSON"

    # ── TÍTULO ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:R1")
    c = ws["A1"]
    c.value = "COMISSÃO WILSON  ·  QUADRATA SEGUROS  |  Vigência: 23/04/2025 – 23/04/2026"
    c.font  = Font(name="Calibri", bold=True, size=14, color=COR_BRANCO)
    c.fill  = fill(COR_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── LEGENDA STATUS ───────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:R2")
    leg = ws["A2"]
    leg.value = (
        "🟢 Ativo / Adimplente   🔴 Inativo   🟡 Inadimplente   "
        "— Use os filtros (▼) na coluna STATUS e ADIMPLÊNCIA para filtrar"
    )
    leg.font      = Font(name="Calibri", italic=True, size=9, color=COR_BRANCO)
    leg.fill      = fill("2E75B6")
    leg.alignment = Alignment(horizontal="center", vertical="center")

    current_row   = 3          # próxima linha livre
    all_data_rows = []         # (row_num, has_data) para validação

    for mes, linhas in MESES.items():
        # ── CABEÇALHO DO MÊS ────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 22
        # Célula do mês (col A)
        mc = ws.cell(row=current_row, column=1, value=mes)
        mc.font      = Font(name="Calibri", bold=True, size=10, color=COR_BRANCO)
        mc.fill      = fill(COR_MES)
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border    = BORDA_MES

        # Demais cabeçalhos (col B..R)
        for ci, hdr in enumerate(DATA_HEADERS, 2):
            hc = ws.cell(row=current_row, column=ci, value=hdr)
            hc.font      = Font(name="Calibri", bold=True, size=9, color=COR_BRANCO)
            hc.fill      = fill(COR_MES if hdr else "1A5276")
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border    = BORDA_MES

        header_row = current_row
        current_row += 1

        if not linhas:
            # Mês sem dados → linha de aviso
            ws.row_dimensions[current_row].height = 16
            ws.merge_cells(f"A{current_row}:R{current_row}")
            ac = ws[f"A{current_row}"]
            ac.value     = f"⚠  Dados de {mes} ainda não inseridos — preencher manualmente"
            ac.font      = Font(name="Calibri", italic=True, size=9, color="7F6000")
            ac.fill      = fill(COR_AMAR_CLR)
            ac.alignment = Alignment(horizontal="center")
            current_row += 1
            # linha em branco
            current_row += 1
            continue

        total_receber = 0.0
        for i, d in enumerate(linhas):
            (cliente, seg, item, premio, cpct,
             com, p26, p2da, recebido,
             vezes, receber, pnum, ptot) = d

            ws.row_dimensions[current_row].height = 16
            alt = COR_CINZA_CLR if i % 2 == 0 else COR_BRANCO

            def wr(col, val, fmt=None, bold=False, center=False, cor_txt=None):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font   = Font(name="Calibri", size=9,
                                   bold=bold, color=cor_txt or "000000")
                cell.fill   = fill(alt)
                cell.border = BORDA_FINA
                cell.alignment = Alignment(
                    horizontal="center" if center else
                    ("right" if fmt else "left"),
                    vertical="center")
                if fmt:
                    cell.number_format = fmt
                return cell

            wr(1,  cliente,  bold=True)
            wr(2,  premio,   fmt='#,##0.00', cor_txt="1F3864")
            wr(3,  seg,      center=True)
            wr(4,  item,     center=True)
            wr(5,  cpct,     fmt='0"%"',    center=True)
            wr(6,  com,      fmt='#,##0.00')
            wr(7,  p26,      fmt='#,##0.00')
            wr(8,  p2da,     fmt='#,##0.00')
            wr(9,  recebido, fmt='#,##0.00')
            wr(10, vezes,    center=True)
            wr(11, receber,  fmt='#,##0.00')

            # PARCELA: num | DE | total
            last = (pnum == ptot)
            cor_parc = "C00000" if last else "1F3864"
            wr(12, pnum,  bold=True, center=True, cor_txt=cor_parc)
            wr(13, "DE",  center=True)
            wr(14, ptot,  bold=True, center=True)

            # AJUSTE (vazio – preencher se necessário)
            wr(15, None, fmt='#,##0.00')

            # STATUS
            sc = ws.cell(row=current_row, column=16, value="Ativo")
            sc.font      = Font(name="Calibri", bold=True, size=9, color="375623")
            sc.fill      = fill(COR_VERDE_CLR)
            sc.border    = BORDA_FINA
            sc.alignment = Alignment(horizontal="center", vertical="center")

            # ADIMPLÊNCIA
            ac2 = ws.cell(row=current_row, column=17, value="Adimplente")
            ac2.font      = Font(name="Calibri", bold=True, size=9, color="375623")
            ac2.fill      = fill(COR_VERDE_CLR)
            ac2.border    = BORDA_FINA
            ac2.alignment = Alignment(horizontal="center", vertical="center")

            # OBS
            oc = ws.cell(row=current_row, column=18, value="")
            oc.fill   = fill(alt)
            oc.border = BORDA_FINA

            all_data_rows.append(current_row)
            total_receber += receber
            current_row += 1

        # ── TOTAL DO MÊS ────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 18
        ws.merge_cells(f"A{current_row}:J{current_row}")
        tc = ws[f"A{current_row}"]
        tc.value     = f"TOTAL  {mes.upper()}"
        tc.font      = Font(name="Calibri", bold=True, size=9, color=COR_BRANCO)
        tc.fill      = fill(COR_TITULO)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border    = BORDA_FINA

        tv = ws.cell(row=current_row, column=11, value=total_receber)
        tv.font          = Font(name="Calibri", bold=True, size=9, color=COR_BRANCO)
        tv.fill          = fill(COR_TITULO)
        tv.number_format = 'R$ #,##0.00'
        tv.alignment     = Alignment(horizontal="right", vertical="center")
        tv.border        = BORDA_FINA
        for ci in range(12, 19):
            ec = ws.cell(row=current_row, column=ci)
            ec.fill   = fill(COR_TITULO)
            ec.border = BORDA_FINA

        current_row += 2   # linha em branco entre meses

    # ── VALIDAÇÃO (dropdowns STATUS e ADIMPLÊNCIA) ───────────────────────────
    if all_data_rows:
        first_dr = all_data_rows[0]
        last_dr  = all_data_rows[-1]

        dv_s = DataValidation(
            type="list", formula1='"Ativo,Inativo"',
            allow_blank=False, showErrorMessage=True,
            errorTitle="Inválido", error="Escolha: Ativo ou Inativo")
        dv_s.sqref = f"P{first_dr}:P{last_dr}"
        ws.add_data_validation(dv_s)

        dv_a = DataValidation(
            type="list", formula1='"Adimplente,Inadimplente"',
            allow_blank=False, showErrorMessage=True,
            errorTitle="Inválido", error="Escolha: Adimplente ou Inadimplente")
        dv_a.sqref = f"Q{first_dr}:Q{last_dr}"
        ws.add_data_validation(dv_a)

        full_range = f"A{first_dr}:R{last_dr}"

        # Inativo → vermelho claro
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[f'$P{first_dr}="Inativo"'],
            fill=fill(COR_VERM_CLR),
            font=Font(name="Calibri", size=9, color="C00000")))

        # Inadimplente → amarelo
        ws.conditional_formatting.add(full_range, FormulaRule(
            formula=[f'$Q{first_dr}="Inadimplente"'],
            fill=fill(COR_AMAR_CLR),
            font=Font(name="Calibri", size=9, color="7F6000")))

    # ── LARGURAS ────────────────────────────────────────────────────────────
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
def build_resumo(wb):
    ws = wb.create_sheet("RESUMO MENSAL")

    ws.row_dimensions[1].height = 30
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "RESUMO MENSAL — COMISSÕES WILSON  |  Vigência 23/04/2025 – 23/04/2026"
    c.font      = Font(name="Calibri", bold=True, size=13, color=COR_BRANCO)
    c.fill      = fill(COR_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["MÊS","Nº CLIENTES","TOTAL PRÊMIOS","TOTAL COM.","TOTAL RECEBER","SITUAÇÃO","OBS"]
    ws.row_dimensions[2].height = 20
    for ci, h in enumerate(headers, 1):
        hc = ws.cell(row=2, column=ci, value=h)
        hc.font      = Font(name="Calibri", bold=True, size=10, color=COR_BRANCO)
        hc.fill      = fill(COR_MES)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border    = BORDA_FINA

    totals = {"premio":0, "com":0, "receber":0}
    for ri, (mes, linhas) in enumerate(MESES.items(), 3):
        ws.row_dimensions[ri].height = 18
        alt = COR_CINZA_CLR if ri % 2 == 0 else COR_BRANCO

        if linhas:
            n   = len(linhas)
            pr  = sum(r[3] for r in linhas)
            cm  = sum(r[5] for r in linhas)
            rc  = sum(r[11] for r in linhas)
            sit = "✅ Dados OK"
            totals["premio"]  += pr
            totals["com"]     += cm
            totals["receber"] += rc
        else:
            n, pr, cm, rc = 0, None, None, None
            sit = "⚠ Aguardando dados"

        row = [mes, n or "—", pr, cm, rc, sit, ""]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = fill(alt)
            cell.border = BORDA_FINA
            cell.font   = Font(name="Calibri", size=10)
            cell.alignment = Alignment(
                horizontal="center" if ci in (1,2,6,7) else "right",
                vertical="center")
            if ci in (3,4,5) and val is not None:
                cell.number_format = 'R$ #,##0.00'
            if ci == 6 and "Aguardando" in str(val):
                cell.font = Font(name="Calibri", size=10,
                                 color="7F6000", italic=True)

    # Linha de total
    tr = len(MESES) + 3
    ws.row_dimensions[tr].height = 22
    for ci in range(1, 8):
        tc = ws.cell(row=tr, column=ci)
        tc.fill   = fill(COR_TITULO)
        tc.border = BORDA_FINA
        tc.font   = Font(name="Calibri", bold=True, color=COR_BRANCO, size=10)
        tc.alignment = Alignment(horizontal="center" if ci!=1 else "right",
                                  vertical="center")
    ws.cell(row=tr, column=1).value = "TOTAL"
    ws.cell(row=tr, column=3).value = totals["premio"]
    ws.cell(row=tr, column=4).value = totals["com"]
    ws.cell(row=tr, column=5).value = totals["receber"]
    for ci in (3,4,5):
        ws.cell(row=tr, column=ci).number_format = 'R$ #,##0.00'
        ws.cell(row=tr, column=ci).alignment = Alignment(horizontal="right")

    for ci, w in enumerate([12,14,18,16,16,18,20], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    wb = openpyxl.Workbook()
    build_sheet(wb)
    build_resumo(wb)
    out = "/home/user/whatsapp-webhook/comissoes_wilson_2025_26.xlsx"
    wb.save(out)
    print(f"✅  Planilha salva em: {out}")
